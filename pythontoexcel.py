import openpyxl as xl
from openpyxl.styles import Font

#create a new excel workbook

wb = xl.workbook()

ws = wb.active

ws.title = 'First Sheet'

#create a new worksheet

wb.create_sheet(index=1,title= 'second sheet')

#sav workbook 

wb.save('Pythontoexcel.xlsx')

ws['A1'] = "Invoice"

fontobj = Font(name= "Times New Roman", size=24, bold=True)

ws['A1'].font = fontobj

ws['A2'].font = "Tires"
ws['A3'].font = "brakes"
ws['A4'].font = "Alignment"

ws['B2'].font = 450
ws['B3'].font = 225.50
ws['B4'].font = 150

ws['A8'] = "Total"
ws['A8'].font = fontobj

ws.column_dimensions['A'].width = 25

ws.merge_cells('A1:B1')

ws['B8'] = '=SUM(B2:B7)'


#new for prduce 

write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook('ProduceReport.xlsx')
read.ws = read_wb["ProducerReport"]
